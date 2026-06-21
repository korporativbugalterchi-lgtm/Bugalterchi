import logging
import asyncio
import json
import os
from datetime import datetime, timedelta
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import CommandStart, Command
from aiogram.types import (InlineKeyboardMarkup, InlineKeyboardButton,
                           ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove,
                           FSInputFile)

# ==================== SOZLAMALAR ====================
BOT_TOKEN = "8326552518:AAGFiqzMTHwj-o65oUZ9UV_h63OgCXk0w6Q"
ADMIN_ID = 422549832
CHANNEL_LINK = "https://t.me/your_channel"
CHANNEL_ID = -1001234567890

USERS_FILE = "users.json"
SCHEDULE_FILE = "schedule.json"
POSTS_FILE = "posts.json"
TIMED_FILE = "timed_posts.json"
FORWARD_MAP_FILE = "forward_map.json"
BLOCKED_FILE = "blocked.json"
HISTORY_FILE = "history.json"

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()

# ==================== FAYL FUNKSIYALARI ====================

def load_json(filename):
    if os.path.exists(filename):
        with open(filename, "r", encoding="utf-8-sig") as f:
            return json.load(f)
    return {}

def save_json(filename, data):
    with open(filename, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def format_delay(delay_hours: float) -> str:
    if delay_hours == 0:
        return "darhol"
    if delay_hours < 1:
        minutes = round(delay_hours * 60)
        return f"{minutes} minut"
    if delay_hours < 24:
        return f"{int(delay_hours)} soat"
    return f"{int(delay_hours // 24)} kun"

def feedback_keyboard(ref: str):
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="✅ Ha", callback_data=f"fb_yes_{ref}"),
            InlineKeyboardButton(text="❌ Yo'q", callback_data=f"fb_no_{ref}"),
        ]
    ])

def next_step_keyboard():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="➕ Yangi post qo'shish", callback_data="add_new_post")],
        [InlineKeyboardButton(text="🎬 Ko'rdingizmi? xabarini yuborish", callback_data="send_feedback_now")],
        [InlineKeyboardButton(text="📋 Postlar ro'yxati", callback_data="set_scheduled")],
        [InlineKeyboardButton(text="🏠 Bosh menyu", callback_data="back_admin")],
    ])

# ==================== BLOKLASH TEKSHIRISH ====================

def is_blocked(user_id: int) -> bool:
    blocked = load_json(BLOCKED_FILE)
    return str(user_id) in blocked

# ==================== OBUNA TEKSHIRISH ====================

async def check_subscription(user_id: int) -> bool:
    try:
        member = await bot.get_chat_member(CHANNEL_ID, user_id)
        return member.status not in ["left", "kicked", "banned"]
    except:
        return True

def sub_keyboard():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📢 Kanalga obuna bo'lish", url=CHANNEL_LINK)],
        [InlineKeyboardButton(text="✅ Obuna bo'ldim", callback_data="check_sub")]
    ])

def phone_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="📱 Telefon raqamimni ulashish", request_contact=True)]],
        resize_keyboard=True,
        one_time_keyboard=True
    )

# ==================== XUSH KELIBSIZ ====================

async def send_welcome(user_id: int, name: str):
    posts = load_json(POSTS_FILE)
    welcome = posts.get("welcome", {})
    if welcome.get("photo_id"):
        await bot.send_photo(user_id, photo=welcome["photo_id"], caption=welcome.get("text", ""), parse_mode="HTML")
    elif welcome.get("text"):
        await bot.send_message(user_id, welcome["text"], parse_mode="HTML")
    else:
        await bot.send_message(user_id,
            f"👋 Assalomu alaykum, <b>{name}</b>!\n\nMehriddin Nuriddinov Botga xush kelibsiz! 🎉\n\n"
            f"📚 Tizimli biznes qurish, xodimlarni tizimli boshqarish, marketing va sotuv bo'yicha tajribadan kelib chiqqan xulosa va biznesda o'qiganlarim bo'yicha instrumentlarni berib boraman. \n\n"
            f"🔔 Shu va shunga o‘xshash foydali ma'lumotlar sizni qiziqtirsa, biz bilan birga bo‘ling.", parse_mode="HTML")
    if welcome.get("voice_id"):
        await bot.send_voice(user_id, voice=welcome["voice_id"])

# ==================== START ====================

@dp.message(CommandStart())
async def start_handler(message: types.Message):
    if is_blocked(message.from_user.id):
        await message.answer("⛔ Siz botdan bloklangansiz.")
        return
    users = load_json(USERS_FILE)
    uid = str(message.from_user.id)
    is_sub = await check_subscription(message.from_user.id)
    if not is_sub:
        await message.answer("👋 Salom!\n\nBotdan foydalanish uchun avval kanalga obuna bo'ling! 👇", reply_markup=sub_keyboard())
        return
    if uid in users and users[uid].get("phone"):
        await send_welcome(message.from_user.id, message.from_user.first_name or "Foydalanuvchi")
        return
    if uid not in users:
        users[uid] = {
            "id": message.from_user.id,
            "first_name": message.from_user.first_name or "",
            "last_name": message.from_user.last_name or "",
            "username": message.from_user.username or "",
            "phone": "",
            "joined": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "left": "",
        }
        save_json(USERS_FILE, users)
    await message.answer(
        f"👋 Salom, <b>{message.from_user.first_name}</b>!\n\nDavom etish uchun telefon raqamingizni ulashing 👇",
        parse_mode="HTML", reply_markup=phone_keyboard())

# ==================== TELEFON ====================

@dp.message(F.contact)
async def contact_handler(message: types.Message):
    if is_blocked(message.from_user.id):
        return
    users = load_json(USERS_FILE)
    uid = str(message.from_user.id)
    is_new = uid not in users or not users[uid].get("phone")
    if uid in users:
        users[uid]["phone"] = message.contact.phone_number
        users[uid]["first_name"] = message.from_user.first_name or ""
        users[uid]["last_name"] = message.from_user.last_name or ""
        users[uid]["username"] = message.from_user.username or ""
        was_left = bool(users[uid].get("left"))
        users[uid]["joined"] = datetime.now().strftime("%Y-%m-%d %H:%M")
        users[uid]["left"] = ""
        save_json(USERS_FILE, users)
        reason = "rejoined" if was_left else "joined"
        save_history(uid, users[uid], reason=reason, by="self")
    await message.answer("✅ Rahmat!", reply_markup=ReplyKeyboardRemove())
    await send_welcome(message.from_user.id, message.from_user.first_name or "Foydalanuvchi")
    await schedule_posts_for_user(message.from_user.id)

# ==================== REJALIK POSTLAR ====================

async def schedule_posts_for_user(user_id: int):
    posts = load_json(POSTS_FILE)
    schedule = load_json(SCHEDULE_FILE)
    uid = str(user_id)
    now = datetime.now()
    schedule[uid] = []
    post_list = []
    for key, post in posts.items():
        if key.startswith("post_") and post.get("delay_hours") is not None:
            post_list.append((key, post))
    post_list.sort(key=lambda x: int(x[0].replace("post_", "")))
    cursor = now
    for key, post in post_list:
        delay_hours = post["delay_hours"]
        cursor = cursor + timedelta(hours=delay_hours)
        schedule[uid].append({
            "post_key": key,
            "send_time": cursor.strftime("%Y-%m-%d %H:%M"),
            "sent": False
        })
    save_json(SCHEDULE_FILE, schedule)

# ==================== SCHEDULER ====================

async def check_and_send_scheduled():
    while True:
        try:
            schedule = load_json(SCHEDULE_FILE)
            posts = load_json(POSTS_FILE)
            now = datetime.now()
            changed = False
            for uid, user_schedule in schedule.items():
                if is_blocked(int(uid)):
                    continue
                for item in user_schedule:
                    if item["sent"]:
                        continue
                    send_time = datetime.strptime(item["send_time"], "%Y-%m-%d %H:%M")
                    if now >= send_time:
                        post = posts.get(item["post_key"])
                        if post:
                            try:
                                if post.get("poll"):
                                    p = post["poll"]
                                    await bot.send_poll(int(uid), question=p["question"], options=p["options"], is_anonymous=p.get("is_anonymous", True))
                                elif post.get("photo_id"):
                                    await bot.send_photo(int(uid), photo=post["photo_id"], caption=post.get("text", ""), parse_mode="HTML")
                                elif post.get("video_id"):
                                    await bot.send_video(int(uid), video=post["video_id"], caption=post.get("text", ""), parse_mode="HTML")
                                elif post.get("video_note_id"):
                                    await bot.send_video_note(int(uid), video_note=post["video_note_id"])
                                elif post.get("text"):
                                    await bot.send_message(int(uid), post["text"], parse_mode="HTML")
                                if post.get("voice_id"):
                                    await bot.send_voice(int(uid), voice=post["voice_id"])
                            except Exception as e:
                                logger.error(f"Scheduled send error {uid}: {e}")
                        item["sent"] = True
                        changed = True
            if changed:
                save_json(SCHEDULE_FILE, schedule)

            timed = load_json(TIMED_FILE)
            timed_changed = False
            for pid, p in list(timed.items()):
                if p.get("sent"):
                    continue
                send_time = datetime.strptime(p["send_time"], "%Y-%m-%d %H:%M")
                if now >= send_time:
                    users = load_json(USERS_FILE)
                    # Feedback (Ko'rdingizmi?) xabari
                    if p.get("feedback"):
                        ref = p.get("ref", f"auto_{pid}")
                        for uid2 in users.keys():
                            if is_blocked(int(uid2)):
                                continue
                            try:
                                await bot.send_message(int(uid2), "🎬 Ko'rdingizmi?",
                                    reply_markup=feedback_keyboard(ref))
                                await asyncio.sleep(0.05)
                            except:
                                pass
                        timed[pid]["sent"] = True
                        timed_changed = True
                        continue

                    for uid2 in users.keys():
                        if is_blocked(int(uid2)):
                            continue
                        try:
                            if p.get("poll"):
                                pl = p["poll"]
                                await bot.send_poll(int(uid2), question=pl["question"], options=pl["options"], is_anonymous=pl.get("is_anonymous", True))
                            elif p.get("photo_id"):
                                await bot.send_photo(int(uid2), photo=p["photo_id"], caption=p.get("text", ""), parse_mode="HTML")
                            elif p.get("video_id"):
                                await bot.send_video(int(uid2), video=p["video_id"], caption=p.get("text", ""), parse_mode="HTML")
                            elif p.get("video_note_id"):
                                await bot.send_video_note(int(uid2), video_note=p["video_note_id"])
                            elif p.get("text"):
                                await bot.send_message(int(uid2), p["text"], parse_mode="HTML")
                            if p.get("voice_id"):
                                await bot.send_voice(int(uid2), voice=p["voice_id"])
                            await asyncio.sleep(0.05)
                        except:
                            pass
                    timed[pid]["sent"] = True
                    timed_changed = True
            if timed_changed:
                save_json(TIMED_FILE, timed)
        except Exception as e:
            logger.error(f"Scheduler error: {e}")
        await asyncio.sleep(60)

# ==================== FEEDBACK ====================

@dp.callback_query(F.data.startswith("fb_yes_") | F.data.startswith("fb_no_"))
async def feedback_cb(callback: types.CallbackQuery):
    is_yes = callback.data.startswith("fb_yes_")
    ref = callback.data.replace("fb_yes_", "").replace("fb_no_", "")
    user = callback.from_user
    name = f"{user.first_name or ''} {user.last_name or ''}".strip() or "Nomsiz"
    username = f"@{user.username}" if user.username else "—"
    answer_text = "✅ Ha" if is_yes else "❌ Yo'q"
    try:
        await callback.message.edit_reply_markup(reply_markup=None)
    except:
        pass
    await callback.answer("Rahmat, javobingiz qabul qilindi!")
    await callback.message.answer("✅ Rahmat, javobingiz yuborildi!")
    try:
        sent_msg = await bot.send_message(
            ADMIN_ID,
            f"🎬 <b>Post:</b> {ref}\n"
            f"👤 <b>{name}</b> ({username})\n"
            f"🆔 <code>{user.id}</code>\n"
            f"📝 Javob: <b>{answer_text}</b>\n\n"
            f"✍️ Foydalanuvchiga javob yozish uchun shu xabarga <b>Reply</b> qiling.",
            parse_mode="HTML"
        )
        forward_map = load_json(FORWARD_MAP_FILE)
        forward_map[str(sent_msg.message_id)] = user.id
        save_json(FORWARD_MAP_FILE, forward_map)
    except Exception as e:
        logger.error(f"Feedback notify admin error: {e}")

# ==================== OBUNA CALLBACK ====================

@dp.callback_query(F.data == "check_sub")
async def check_sub_cb(callback: types.CallbackQuery):
    is_sub = await check_subscription(callback.from_user.id)
    if is_sub:
        await callback.message.delete()
        users = load_json(USERS_FILE)
        uid = str(callback.from_user.id)
        if uid in users and users[uid].get("phone"):
            await send_welcome(callback.from_user.id, callback.from_user.first_name or "")
        else:
            await callback.message.answer("✅ Obuna tasdiqlandi!\n\nEndi telefon raqamingizni ulashing 👇", reply_markup=phone_keyboard())
    else:
        await callback.answer("❌ Hali obuna bo'lmadingiz!", show_alert=True)

# ==================== ADMIN PANEL ====================

admin_state = {}

def admin_main_keyboard(users):
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=f"👥 A'zolar ro'yxati ({len(users)} ta)", callback_data="show_users")],
        [InlineKeyboardButton(text="📊 Statistika", callback_data="show_stats")],
        [InlineKeyboardButton(text="📋 Foydalanuvchilar tarixi", callback_data="show_all_history")],
        [InlineKeyboardButton(text="📊 Excel eksport", callback_data="export_excel")],
        [InlineKeyboardButton(text="➕ Foydalanuvchi qo'shish", callback_data="add_user")],
        [InlineKeyboardButton(text="📢 Oddiy post (hozir)", callback_data="broadcast_now")],
        [InlineKeyboardButton(text="⏰ Rejalik post (aniq vaqt)", callback_data="broadcast_timed")],
        [InlineKeyboardButton(text="✏️ Xush kelibsiz xabarini sozlash", callback_data="set_welcome")],
        [InlineKeyboardButton(text="📋 Rejalik postlarni sozlash", callback_data="set_scheduled")],
        [InlineKeyboardButton(text="🎬 Ko'rdingizmi? xabarini yuborish", callback_data="send_feedback_now")],
    ])

@dp.message(Command("admin"))
async def admin_panel(message: types.Message):
    if message.from_user.id != ADMIN_ID:
        return
    users = load_json(USERS_FILE)
    await message.answer(
        f"🔧 <b>Admin Panel</b>\n👥 Jami a'zolar: <b>{len(users)} ta</b>",
        reply_markup=admin_main_keyboard(users), parse_mode="HTML")

@dp.callback_query(F.data == "back_admin")
async def back_admin(callback: types.CallbackQuery):
    if callback.from_user.id != ADMIN_ID:
        return
    users = load_json(USERS_FILE)
    await callback.message.answer(
        f"🔧 <b>Admin Panel</b>\n👥 Jami a'zolar: <b>{len(users)} ta</b>",
        reply_markup=admin_main_keyboard(users), parse_mode="HTML")
    await callback.answer()

# ==================== STATISTIKA ====================

MONTHS_UZ = ["Yanvar","Fevral","Mart","Aprel","May","Iyun","Iyul","Avgust","Sentabr","Oktabr","Noyabr","Dekabr"]
DAYS_UZ = ["Du","Se","Ch","Pa","Ju","Sh","Ya"]

def calendar_keyboard(year: int, month: int, prefix: str):
    """Kalendar inline klaviatura — prefix: 'from' yoki 'to'"""
    import calendar
    buttons = []
    # Sarlavha: Oy nomi + Yil
    month_name = MONTHS_UZ[month - 1]
    buttons.append([InlineKeyboardButton(
        text=f"◀️", callback_data=f"cal_{prefix}_prev_{year}_{month}"),
        InlineKeyboardButton(text=f"📅 {month_name} {year}", callback_data="cal_ignore"),
        InlineKeyboardButton(text=f"▶️", callback_data=f"cal_{prefix}_next_{year}_{month}"),
    ])
    # Hafta kunlari
    buttons.append([InlineKeyboardButton(text=d, callback_data="cal_ignore") for d in DAYS_UZ])
    # Kunlar
    cal = calendar.monthcalendar(year, month)
    for week in cal:
        row = []
        for day in week:
            if day == 0:
                row.append(InlineKeyboardButton(text=" ", callback_data="cal_ignore"))
            else:
                row.append(InlineKeyboardButton(
                    text=str(day),
                    callback_data=f"cal_{prefix}_day_{year}_{month}_{day}"
                ))
        buttons.append(row)
    buttons.append([InlineKeyboardButton(text="🔙 Bekor qilish", callback_data="back_admin")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)

def compute_stats(users, blocked, from_date=None, to_date=None):
    """Statistikani hisoblash. from_date/to_date — date ob'ektlar."""
    new_in_range = 0
    left_in_range = 0
    left_total = 0
    active = 0
    total = len(users)
    for u in users.values():
        joined_str = u.get("joined", "")
        left_str = u.get("left", "")
        if left_str:
            left_total += 1
        else:
            active += 1
        if joined_str and from_date and to_date:
            try:
                jd = datetime.strptime(joined_str, "%Y-%m-%d %H:%M").date()
                if from_date <= jd <= to_date:
                    new_in_range += 1
            except: pass
        if left_str and from_date and to_date:
            try:
                ld = datetime.strptime(left_str, "%Y-%m-%d %H:%M").date()
                if from_date <= ld <= to_date:
                    left_in_range += 1
            except: pass
    return total, active, left_total, new_in_range, left_in_range

def stats_text(users, blocked, from_date=None, to_date=None):
    total, active, left_total, new_in_range, left_in_range = compute_stats(users, blocked, from_date, to_date)
    now = datetime.now()
    if from_date and to_date:
        period = f"📆 <b>Davr:</b> {from_date.strftime('%d.%m.%Y')} — {to_date.strftime('%d.%m.%Y')}"
    else:
        period = f"📆 <b>Davr:</b> Barcha vaqt"

    bar_active = "🟢" * min(active, 10) if active else "⬜"
    bar_left = "🔴" * min(left_total, 10) if left_total else "⬜"

    text = (
        f"╔══════════════════╗\n"
        f"║  📊  BOT STATISTIKASI  ║\n"
        f"╚══════════════════╝\n\n"
        f"{period}\n"
        f"🕐 <b>Yangilangan:</b> {now.strftime('%d.%m.%Y %H:%M')}\n\n"
        f"┌─────────────────────\n"
        f"│ 👥 <b>Jami a'zolar:</b>  <b>{total}</b> ta\n"
        f"│ ✅ <b>Faol:</b>  <b>{active}</b> ta  {bar_active}\n"
        f"│ 🚪 <b>Chiqib ketgan:</b>  <b>{left_total}</b> ta  {bar_left}\n"
        f"│ 🚫 <b>Bloklangan:</b>  <b>{len(blocked)}</b> ta\n"
        f"└─────────────────────\n\n"
    )
    if from_date and to_date:
        text += (
            f"┌─────────────────────\n"
            f"│ 🆕 <b>Yangi qo'shildi:</b>  <b>{new_in_range}</b> ta\n"
            f"│ 🚶 <b>Chiqib ketdi:</b>  <b>{left_in_range}</b> ta\n"
            f"└─────────────────────\n"
        )
    return text

@dp.callback_query(F.data == "show_stats")
async def show_stats(callback: types.CallbackQuery):
    if callback.from_user.id != ADMIN_ID:
        return
    users = load_json(USERS_FILE)
    blocked = load_json(BLOCKED_FILE)
    now = datetime.now()
    # Umumiy statistika + kalendar tanlash tugmasi
    text = stats_text(users, blocked)
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📅 Davr bo'yicha ko'rish", callback_data="stats_pick_from")],
        [InlineKeyboardButton(text="🔙 Admin panel", callback_data="back_admin")],
    ])
    await callback.message.answer(text, parse_mode="HTML", reply_markup=kb)
    await callback.answer()

# ---- Davr tanlash: FROM ----
@dp.callback_query(F.data == "stats_pick_from")
async def stats_pick_from(callback: types.CallbackQuery):
    if callback.from_user.id != ADMIN_ID:
        return
    now = datetime.now()
    await callback.message.answer(
        "📅 <b>Boshlanish sanasini tanlang:</b>",
        parse_mode="HTML",
        reply_markup=calendar_keyboard(now.year, now.month, "from")
    )
    await callback.answer()

# ---- Kalendar: oldingi/keyingi oy ----
@dp.callback_query(F.data.startswith("cal_from_prev_") | F.data.startswith("cal_from_next_") |
                   F.data.startswith("cal_to_prev_") | F.data.startswith("cal_to_next_"))
async def cal_nav(callback: types.CallbackQuery):
    if callback.from_user.id != ADMIN_ID:
        return
    parts = callback.data.split("_")
    # cal_{prefix}_{prev/next}_{year}_{month}
    prefix = parts[1]
    direction = parts[2]
    year = int(parts[3])
    month = int(parts[4])
    if direction == "prev":
        month -= 1
        if month < 1:
            month = 12
            year -= 1
    else:
        month += 1
        if month > 12:
            month = 1
            year += 1
    label = "Boshlanish" if prefix == "from" else "Tugash"
    try:
        await callback.message.edit_reply_markup(reply_markup=calendar_keyboard(year, month, prefix))
    except:
        await callback.message.answer(
            f"📅 <b>{label} sanasini tanlang:</b>",
            parse_mode="HTML",
            reply_markup=calendar_keyboard(year, month, prefix))
    await callback.answer()

# ---- Kalendar: kun tanlash ----
@dp.callback_query(F.data.startswith("cal_from_day_") | F.data.startswith("cal_to_day_"))
async def cal_day(callback: types.CallbackQuery):
    if callback.from_user.id != ADMIN_ID:
        return
    parts = callback.data.split("_")
    # cal_{prefix}_day_{year}_{month}_{day}
    prefix = parts[1]
    year = int(parts[3])
    month = int(parts[4])
    day = int(parts[5])
    from datetime import date as date_type
    chosen = date_type(year, month, day)

    if prefix == "from":
        admin_state[ADMIN_ID] = admin_state.get(ADMIN_ID, {})
        admin_state[ADMIN_ID]["stats_from"] = chosen.strftime("%Y-%m-%d")
        # Tugash sanasi uchun kalendar
        await callback.message.answer(
            f"✅ Boshlanish: <b>{chosen.strftime('%d.%m.%Y')}</b>\n\n"
            f"📅 <b>Tugash sanasini tanlang:</b>",
            parse_mode="HTML",
            reply_markup=calendar_keyboard(year, month, "to")
        )
    elif prefix == "to":
        from_str = admin_state.get(ADMIN_ID, {}).get("stats_from")
        if not from_str:
            await callback.answer("Avval boshlanish sanasini tanlang!", show_alert=True)
            return
        from_date = datetime.strptime(from_str, "%Y-%m-%d").date()
        to_date = chosen
        if to_date < from_date:
            await callback.answer("❌ Tugash sanasi boshlanishdan oldin bo'lmasin!", show_alert=True)
            return
        users = load_json(USERS_FILE)
        blocked = load_json(BLOCKED_FILE)
        text = stats_text(users, blocked, from_date, to_date)
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📅 Boshqa davr", callback_data="stats_pick_from")],
            [InlineKeyboardButton(text="🔙 Admin panel", callback_data="back_admin")],
        ])
        await callback.message.answer(text, parse_mode="HTML", reply_markup=kb)
        admin_state.get(ADMIN_ID, {}).pop("stats_from", None)

    await callback.answer()

@dp.callback_query(F.data == "cal_ignore")
async def cal_ignore(callback: types.CallbackQuery):
    await callback.answer()

# ==================== A'ZOLAR RO'YXATI (ID BILAN) ====================

@dp.callback_query(F.data == "show_users")
async def show_users(callback: types.CallbackQuery):
    if callback.from_user.id != ADMIN_ID:
        return
    users = load_json(USERS_FILE)
    blocked = load_json(BLOCKED_FILE)
    if not users:
        await callback.answer("Hozircha a'zo yo'q!", show_alert=True)
        return
    text = "👥 <b>A'zolar ro'yxati:</b>\n\n"
    for i, (uid, u) in enumerate(users.items(), 1):
        name = f"{u.get('first_name','')} {u.get('last_name','')}".strip() or "Nomsiz"
        username = f"@{u['username']}" if u.get('username') else "—"
        phone = u.get('phone') or "—"
        joined = u.get('joined', '—')
        left = u.get('left') or "Hali bor"
        block_mark = " 🚫" if uid in blocked else ""
        text += (
            f"{i}.{block_mark} <b>{name}</b>\n"
            f"   🆔 <code>{uid}</code>\n"
            f"   📞 {phone} | 👤 {username}\n"
            f"   📅 {joined} | Chiqdi: {left}\n\n"
        )
    # Boshqarish tugmasi
    manage_btn = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔧 Foydalanuvchini boshqarish", callback_data="manage_user_start")],
        [InlineKeyboardButton(text="🔙 Orqaga", callback_data="back_admin")],
    ])
    for chunk in [text[i:i+4000] for i in range(0, len(text), 4000)]:
        await callback.message.answer(chunk, parse_mode="HTML")
    await callback.message.answer("Quyidagi amallardan birini tanlang:", reply_markup=manage_btn)
    await callback.answer()

# ==================== FOYDALANUVCHI QO'SHISH ====================

@dp.callback_query(F.data == "add_user")
async def add_user_cb(callback: types.CallbackQuery):
    if callback.from_user.id != ADMIN_ID:
        return
    admin_state[ADMIN_ID] = {"action": "add_user_id"}
    await callback.message.answer(
        "➕ <b>Foydalanuvchi qo'shish</b>\n\n"
        "Foydalanuvchining <b>Telegram ID</b> sini yozing:\n"
        "(Masalan: <code>123456789</code>)\n\n"
        "/cancel — bekor qilish",
        parse_mode="HTML"
    )
    await callback.answer()

# ==================== FOYDALANUVCHINI BOSHQARISH ====================

@dp.callback_query(F.data == "manage_user_start")
async def manage_user_start(callback: types.CallbackQuery):
    if callback.from_user.id != ADMIN_ID:
        return
    admin_state[ADMIN_ID] = {"action": "manage_user_id"}
    await callback.message.answer(
        "🔧 <b>Foydalanuvchini boshqarish</b>\n\n"
        "Foydalanuvchining <b>Telegram ID</b> sini yozing:\n"
        "(Ro'yxatdagi <code>ID</code> ni ko'chiring)\n\n"
        "/cancel — bekor qilish",
        parse_mode="HTML"
    )
    await callback.answer()

def user_manage_keyboard(uid: str, is_block: bool):
    block_text = "✅ Blokdan chiqarish" if is_block else "🚫 Bloklash"
    block_cb = f"unblock_user_{uid}" if is_block else f"block_user_{uid}"
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=block_text, callback_data=block_cb)],
        [InlineKeyboardButton(text="🗑 O'chirish (ro'yxatdan)", callback_data=f"remove_user_{uid}")],
        [InlineKeyboardButton(text="📋 Tarixni ko'rish", callback_data=f"show_history_{uid}")],
        [InlineKeyboardButton(text="🔙 Orqaga", callback_data="back_admin")],
    ])

# ==================== BLOKLASH / BLOKDAN CHIQARISH ====================

@dp.callback_query(F.data.startswith("block_user_"))
async def block_user_cb(callback: types.CallbackQuery):
    if callback.from_user.id != ADMIN_ID:
        return
    uid = callback.data.replace("block_user_", "")
    blocked = load_json(BLOCKED_FILE)
    users = load_json(USERS_FILE)
    u = users.get(uid, {})
    name = f"{u.get('first_name','')} {u.get('last_name','')}".strip() or uid
    blocked[uid] = {"blocked_at": datetime.now().strftime("%Y-%m-%d %H:%M")}
    save_json(BLOCKED_FILE, blocked)
    save_history(uid, u, reason="blocked", by="admin")
    try:
        await bot.send_message(int(uid), "⛔ Siz botdan bloklangansiz.")
    except:
        pass
    await callback.message.answer(
        f"🚫 <b>{name}</b> (<code>{uid}</code>) bloklandi!\n"
        f"Endi bu foydalanuvchiga xabar yuborilmaydi.",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ Blokdan chiqarish", callback_data=f"unblock_user_{uid}")],
            [InlineKeyboardButton(text="🔙 Admin panel", callback_data="back_admin")],
        ])
    )
    await callback.answer()

@dp.callback_query(F.data.startswith("unblock_user_"))
async def unblock_user_cb(callback: types.CallbackQuery):
    if callback.from_user.id != ADMIN_ID:
        return
    uid = callback.data.replace("unblock_user_", "")
    blocked = load_json(BLOCKED_FILE)
    users = load_json(USERS_FILE)
    u = users.get(uid, {})
    name = f"{u.get('first_name','')} {u.get('last_name','')}".strip() or uid
    blocked.pop(uid, None)
    save_json(BLOCKED_FILE, blocked)
    save_history(uid, u, reason="unblocked", by="admin")
    try:
        await bot.send_message(int(uid), "✅ Blokingiz olib tashlandi! Botdan yana foydalanishingiz mumkin.")
    except:
        pass
    await callback.message.answer(
        f"✅ <b>{name}</b> (<code>{uid}</code>) blokdan chiqarildi!",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🚫 Qayta bloklash", callback_data=f"block_user_{uid}")],
            [InlineKeyboardButton(text="🔙 Admin panel", callback_data="back_admin")],
        ])
    )
    await callback.answer()

# ==================== RO'YXATDAN O'CHIRISH ====================

def save_history(uid: str, user_data: dict, reason: str, by: str):
    """
    reason: joined | rejoined | self_left | admin_removed | blocked | unblocked
    by:     self | admin | system
    """
    history = load_json(HISTORY_FILE)
    if uid not in history:
        history[uid] = []
    history[uid].append({
        "first_name": user_data.get("first_name", ""),
        "last_name": user_data.get("last_name", ""),
        "username": user_data.get("username", ""),
        "phone": user_data.get("phone", ""),
        "joined": user_data.get("joined", ""),
        "event_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "reason": reason,
        "by": by,
    })
    save_json(HISTORY_FILE, history)

def history_event_text(reason: str) -> str:
    return {
        "joined":        "🟢 Birinchi marta qo'shildi",
        "rejoined":      "🔄 Qayta qo'shildi",
        "self_left":     "🚪 O'zi chiqib ketdi",
        "admin_removed": "🗑 Admin o'chirdi",
        "blocked":       "🚫 Admin blokladi",
        "unblocked":     "✅ Admin blokdan chiqardi",
    }.get(reason, reason)

@dp.callback_query(F.data.startswith("remove_user_"))
async def remove_user_cb(callback: types.CallbackQuery):
    if callback.from_user.id != ADMIN_ID:
        return
    uid = callback.data.replace("remove_user_", "")
    users = load_json(USERS_FILE)
    u = users.pop(uid, {})
    name = f"{u.get('first_name','')} {u.get('last_name','')}".strip() or uid
    save_json(USERS_FILE, users)
    schedule = load_json(SCHEDULE_FILE)
    schedule.pop(uid, None)
    save_json(SCHEDULE_FILE, schedule)
    save_history(uid, u, reason="admin_removed", by="admin")
    await callback.message.answer(
        f"🗑 <b>{name}</b> (<code>{uid}</code>) ro'yxatdan o'chirildi!\n"
        f"(Qayta /start bossa tizimga qo'shiladi)",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📋 Tarix", callback_data=f"show_history_{uid}")],
            [InlineKeyboardButton(text="🔙 Admin panel", callback_data="back_admin")]
        ])
    )
    await callback.answer()

@dp.callback_query(F.data.startswith("show_history_"))
async def show_history_cb(callback: types.CallbackQuery):
    if callback.from_user.id != ADMIN_ID:
        return
    uid = callback.data.replace("show_history_", "")
    history = load_json(HISTORY_FILE)
    users = load_json(USERS_FILE)
    records = history.get(uid, [])

    u = users.get(uid, {})
    name = f"{u.get('first_name','')} {u.get('last_name','')}".strip() or "Nomsiz"
    username = f"@{u['username']}" if u.get('username') else "—"

    if not records:
        await callback.answer("Bu foydalanuvchi tarixi yo'q!", show_alert=True)
        return

    text = (
        f"📋 <b>To'liq tarix</b>\n"
        f"👤 <b>{name}</b> | {username}\n"
        f"🆔 <code>{uid}</code>\n"
        f"━━━━━━━━━━━━━━━━\n\n"
    )
    for i, r in enumerate(records, 1):
        event_text = history_event_text(r.get("reason", ""))
        text += f"{i}. {event_text}\n"
        text += f"   🕐 {r.get('event_at', '—')}\n"
        if r.get("phone") and r["phone"] != "admin_qoshgan":
            text += f"   📞 {r['phone']}\n"
        text += "\n"

    back_kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔙 Orqaga", callback_data="back_admin")]
    ])
    for chunk in [text[i:i+4000] for i in range(0, len(text), 4000)]:
        await callback.message.answer(chunk, parse_mode="HTML", reply_markup=back_kb)
    await callback.answer()

# ---- Barcha tarix ro'yxati ----
@dp.callback_query(F.data == "show_all_history")
async def show_all_history(callback: types.CallbackQuery):
    if callback.from_user.id != ADMIN_ID:
        return
    history = load_json(HISTORY_FILE)
    users = load_json(USERS_FILE)
    if not history:
        await callback.answer("Hozircha tarix yo'q!", show_alert=True)
        return

    # Barcha hodisalarni vaqt bo'yicha saralab chiqarish
    all_events = []
    for uid, records in history.items():
        u = users.get(uid, {})
        name = f"{u.get('first_name','')} {u.get('last_name','')}".strip() or "Nomsiz"
        for r in records:
            all_events.append({
                "uid": uid,
                "name": name,
                "event_at": r.get("event_at", ""),
                "reason": r.get("reason", ""),
                "phone": r.get("phone", ""),
            })
    all_events.sort(key=lambda x: x["event_at"], reverse=True)

    text = f"📋 <b>Barcha tarix</b> (so'nggi {min(len(all_events), 50)} ta)\n━━━━━━━━━━━━━━━━\n\n"
    for e in all_events[:50]:
        event_text = history_event_text(e["reason"])
        text += (
            f"{event_text}\n"
            f"   👤 <b>{e['name']}</b> | 🆔 <code>{e['uid']}</code>\n"
            f"   🕐 {e['event_at']}\n\n"
        )

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔍 ID bo'yicha tarix", callback_data="history_by_id")],
        [InlineKeyboardButton(text="🔙 Admin panel", callback_data="back_admin")],
    ])
    for chunk in [text[i:i+4000] for i in range(0, len(text), 4000)]:
        await callback.message.answer(chunk, parse_mode="HTML")
    await callback.message.answer("Nima qilmoqchisiz?", reply_markup=kb)
    await callback.answer()

@dp.callback_query(F.data == "history_by_id")
async def history_by_id(callback: types.CallbackQuery):
    if callback.from_user.id != ADMIN_ID:
        return
    admin_state[ADMIN_ID] = {"action": "history_search_id"}
    await callback.message.answer(
        "🔍 <b>ID bo'yicha tarix</b>\n\nFoydalanuvchi <b>Telegram ID</b> sini yozing:\n/cancel — bekor qilish",
        parse_mode="HTML"
    )
    await callback.answer()

# ==================== EXCEL ====================

@dp.callback_query(F.data == "export_excel")
async def export_excel(callback: types.CallbackQuery):
    if callback.from_user.id != ADMIN_ID:
        return
    users = load_json(USERS_FILE)
    blocked = load_json(BLOCKED_FILE)
    if not users:
        await callback.answer("Hozircha a'zo yo'q!", show_alert=True)
        return
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "A'zolar"
        headers = ["№", "Ismi", "Familiyasi", "Username", "Telefon", "Telegram ID", "Qo'shilgan", "Chiqgan", "Holati"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        for row, (uid, u) in enumerate(users.items(), 2):
            status = "🚫 Bloklangan" if uid in blocked else ("❌ Chiqib ketgan" if u.get('left') else "✅ Faol")
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=u.get('first_name', ''))
            ws.cell(row=row, column=3, value=u.get('last_name', ''))
            ws.cell(row=row, column=4, value=f"@{u['username']}" if u.get('username') else '')
            ws.cell(row=row, column=5, value=u.get('phone', ''))
            ws.cell(row=row, column=6, value=uid)
            ws.cell(row=row, column=7, value=u.get('joined', ''))
            ws.cell(row=row, column=8, value=u.get('left', ''))
            ws.cell(row=row, column=9, value=status)
        for col in ['A','B','C','D','E','F','G','H','I']:
            ws.column_dimensions[col].width = 18
        filename = f"azolar_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        wb.save(filename)
        await callback.message.answer_document(FSInputFile(filename),
            caption=f"📊 <b>A'zolar ro'yxati</b>\n👥 Jami: {len(users)} ta", parse_mode="HTML")
        os.remove(filename)
    except ImportError:
        await callback.message.answer("❌ <code>pip install openpyxl</code> yozing!", parse_mode="HTML")
    await callback.answer()

# ==================== ODDIY POST ====================

@dp.callback_query(F.data == "broadcast_now")
async def broadcast_now(callback: types.CallbackQuery):
    if callback.from_user.id != ADMIN_ID:
        return
    admin_state[ADMIN_ID] = {"action": "broadcast_now"}
    await callback.message.answer(
        "📢 <b>Hozir yuborish</b>\n\n"
        "Post yuboring (rasm, video, dumaloq video, matn)\n"
        "yoki opros uchun /poll yozing.\n\n"
        "/cancel — bekor qilish",
        parse_mode="HTML")
    await callback.answer()

@dp.callback_query(F.data == "send_feedback_now")
async def send_feedback_now(callback: types.CallbackQuery):
    if callback.from_user.id != ADMIN_ID:
        return
    users = load_json(USERS_FILE)
    if not users:
        await callback.answer("Hozircha a'zo yo'q!", show_alert=True)
        return
    ref = f"manual_{int(datetime.now().timestamp())}"
    success = failed = 0
    await callback.message.answer(f"📤 Yuborilmoqda... ({len(users)} ta a'zo)")
    for uid in users.keys():
        if is_blocked(int(uid)):
            continue
        try:
            await bot.send_message(int(uid), "🎬 Ko'rdingizmi?", reply_markup=feedback_keyboard(ref))
            success += 1
            await asyncio.sleep(0.05)
        except:
            failed += 1
    await callback.message.answer(f"✅ Yuborildi!\n✅ {success} ta | ❌ {failed} ta", reply_markup=next_step_keyboard())
    await callback.answer()

# ==================== REJALIK POST (ANIQ VAQT) ====================

@dp.callback_query(F.data == "broadcast_timed")
async def broadcast_timed(callback: types.CallbackQuery):
    if callback.from_user.id != ADMIN_ID:
        return
    admin_state[ADMIN_ID] = {"action": "timed_wait_post"}
    await callback.message.answer(
        "⏰ <b>Rejalik post (aniq vaqt)</b>\n\n"
        "Postni yuboring (rasm, video, dumaloq video, matn)\n"
        "yoki opros uchun /poll yozing.\n\n"
        "/cancel — bekor qilish",
        parse_mode="HTML")
    await callback.answer()

# ==================== XUSH KELIBSIZ SOZLASH ====================

@dp.callback_query(F.data == "set_welcome")
async def set_welcome(callback: types.CallbackQuery):
    if callback.from_user.id != ADMIN_ID:
        return
    admin_state[ADMIN_ID] = {"action": "set_welcome"}
    await callback.message.answer(
        "✏️ <b>Xush kelibsiz xabarini sozlash</b>\n\nRasm + matn yuboring (yoki faqat matn).\n\n/cancel - bekor qilish",
        parse_mode="HTML")
    await callback.answer()

# ==================== REJALIK POSTLAR SOZLASH ====================

@dp.callback_query(F.data == "set_scheduled")
async def set_scheduled(callback: types.CallbackQuery):
    if callback.from_user.id != ADMIN_ID:
        return
    posts = load_json(POSTS_FILE)
    post_keys = [k for k in posts.keys() if k.startswith("post_")]
    buttons = []
    for key in sorted(post_keys):
        post = posts[key]
        delay = post.get("delay_hours", 0)
        delay_text = format_delay(delay)
        buttons.append([InlineKeyboardButton(text=f"✅ {key} ({delay_text})", callback_data=f"view_post_{key}")])
    buttons.append([InlineKeyboardButton(text="➕ Yangi post qo'shish", callback_data="add_new_post")])
    buttons.append([InlineKeyboardButton(text="🔙 Orqaga", callback_data="back_admin")])
    await callback.message.answer(
        f"📋 <b>Rejalik postlar</b> ({len(post_keys)} ta)",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons), parse_mode="HTML")
    await callback.answer()

@dp.callback_query(F.data.startswith("view_post_"))
async def view_post(callback: types.CallbackQuery):
    if callback.from_user.id != ADMIN_ID:
        return
    post_key = callback.data.replace("view_post_", "")
    posts = load_json(POSTS_FILE)
    post = posts.get(post_key)
    if not post:
        await callback.answer("Post topilmadi!", show_alert=True)
        return
    delay_text = format_delay(post.get("delay_hours", 0))
    await callback.message.answer(f"📌 <b>{post_key} ({delay_text})</b>", parse_mode="HTML")
    if post.get("photo_id"):
        await callback.message.answer_photo(photo=post["photo_id"], caption=post.get("text", "") or "Matn yo'q")
    elif post.get("video_id"):
        await callback.message.answer_video(video=post["video_id"], caption=post.get("text", "") or "Matn yo'q")
    elif post.get("video_note_id"):
        await callback.message.answer_video_note(video_note=post["video_note_id"])
    elif post.get("text"):
        await callback.message.answer(post["text"])
    if post.get("voice_id"):
        await callback.message.answer_voice(voice=post["voice_id"])
    buttons = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✏️ Yangilash", callback_data=f"update_post_{post_key}")],
        [InlineKeyboardButton(text="🗑 O'chirish", callback_data=f"delete_post_{post_key}")],
        [InlineKeyboardButton(text="🔙 Orqaga", callback_data="set_scheduled")],
    ])
    await callback.message.answer("Nima qilmoqchisiz?", reply_markup=buttons)
    await callback.answer()

@dp.callback_query(F.data.startswith("update_post_"))
async def update_post_cb(callback: types.CallbackQuery):
    if callback.from_user.id != ADMIN_ID:
        return
    post_key = callback.data.replace("update_post_", "")
    admin_state[ADMIN_ID] = {"action": "update_post", "post_key": post_key}
    await callback.message.answer(f"✏️ <b>{post_key} yangilash</b>\n\nYangi kontent yuboring:\n/cancel", parse_mode="HTML")
    await callback.answer()

@dp.callback_query(F.data.startswith("delete_post_"))
async def delete_post_cb(callback: types.CallbackQuery):
    if callback.from_user.id != ADMIN_ID:
        return
    post_key = callback.data.replace("delete_post_", "")
    posts = load_json(POSTS_FILE)
    if post_key in posts:
        del posts[post_key]
        save_json(POSTS_FILE, posts)
    await callback.message.answer(f"🗑 <b>{post_key} o'chirildi!</b>", parse_mode="HTML")
    post_keys = [k for k in posts.keys() if k.startswith("post_")]
    buttons = []
    for key in sorted(post_keys):
        p = posts[key]
        delay_text = format_delay(p.get("delay_hours", 0))
        buttons.append([InlineKeyboardButton(text=f"✅ {key} ({delay_text})", callback_data=f"view_post_{key}")])
    buttons.append([InlineKeyboardButton(text="➕ Yangi post qo'shish", callback_data="add_new_post")])
    buttons.append([InlineKeyboardButton(text="🔙 Orqaga", callback_data="back_admin")])
    await callback.message.answer(
        f"📋 <b>Rejalik postlar</b> ({len(post_keys)} ta)",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons), parse_mode="HTML")
    await callback.answer()

@dp.callback_query(F.data == "add_new_post")
async def add_new_post(callback: types.CallbackQuery):
    if callback.from_user.id != ADMIN_ID:
        return
    admin_state[ADMIN_ID] = {"action": "add_post_wait_content"}
    await callback.message.answer(
        "➕ <b>Yangi rejalik post</b>\n\n"
        "Postni yuboring (rasm, video, dumaloq video, matn)\n"
        "yoki opros uchun /poll yozing.\n\n"
        "/cancel — bekor qilish",
        parse_mode="HTML")
    await callback.answer()

# ==================== POLL YUBORISH CALLBACK ====================

@dp.callback_query(F.data == "confirm_poll_send")
async def confirm_poll_send(callback: types.CallbackQuery):
    if callback.from_user.id != ADMIN_ID:
        return
    state = admin_state.get(ADMIN_ID, {})
    action = state.get("action")
    poll_data = state.get("poll_data", {})
    if not poll_data or not poll_data.get("question"):
        await callback.answer("Opros ma'lumoti topilmadi!", show_alert=True)
        return

    if action == "broadcast_now":
        admin_state.pop(ADMIN_ID, None)
        users = load_json(USERS_FILE)
        success = failed = 0
        await callback.message.answer(f"📤 Opros yuborilmoqda... ({len(users)} ta a'zo)")
        for uid in users.keys():
            if is_blocked(int(uid)):
                continue
            try:
                await bot.send_poll(
                    int(uid),
                    question=poll_data["question"],
                    options=poll_data["options"],
                    is_anonymous=poll_data.get("is_anonymous", True)
                )
                success += 1
                await asyncio.sleep(0.05)
            except:
                failed += 1
        await callback.message.answer(f"✅ Opros yuborildi!\n✅ {success} ta | ❌ {failed} ta", reply_markup=next_step_keyboard())

    elif action == "timed_wait_post":
        # Vaqt so'rash
        state["poll_step"] = None
        state["action"] = "timed_wait_time"
        state["post_data"] = {"poll": poll_data}
        admin_state[ADMIN_ID] = state
        await callback.message.answer(
            "⏰ Vaqtni yozing:\n<b>Format: YYYY-MM-DD HH:MM</b>\nMasalan: <code>2025-06-10 14:30</code>",
            parse_mode="HTML")

    elif action == "add_post_wait_content":
        state["poll_step"] = None
        state["action"] = "add_post_wait_delay"
        state["post_data"] = {"poll": poll_data}
        admin_state[ADMIN_ID] = state
        await callback.message.answer(
            "✅ Opros saqlandi!\n\n⏰ Necha vaqt keyin yuborilsin?\n\n"
            "<code>0</code> — darhol\n<code>10 minut</code>\n<code>5 soat</code>\n<code>3 kun</code>",
            parse_mode="HTML")

    await callback.answer()

# ==================== OPROS (POLL) ====================

@dp.message(Command("skip"))
async def skip_cmd(message: types.Message):
    if message.from_user.id != ADMIN_ID:
        return
    state = admin_state.get(ADMIN_ID, {})
    if state.get("action") in ("broadcast_now_feedback_delay", "timed_feedback_delay"):
        admin_state.pop(ADMIN_ID, None)
        await message.answer("⏭ «Ko'rdingizmi?» yuborilmaydi.", reply_markup=next_step_keyboard())
    else:
        await message.answer("⚠️ Hozir skip qilish mumkin emas.")

@dp.message(Command("poll"))
async def poll_cmd(message: types.Message):
    if message.from_user.id != ADMIN_ID:
        return
    state = admin_state.get(ADMIN_ID)
    if not state or state.get("action") not in ("broadcast_now", "timed_wait_post", "add_post_wait_content"):
        await message.answer(
            "⚠️ Avval post yuborish rejimini tanlang:\n"
            "📢 Hozir yuborish, ⏰ Rejalik post yoki 📋 Rejalik postlar sozlash",
            parse_mode="HTML")
        return
    admin_state[ADMIN_ID]["poll_step"] = "wait_question"
    admin_state[ADMIN_ID]["poll_data"] = {}
    await message.answer(
        "📊 <b>Opros yaratish</b>\n\n"
        "1-qadam: Savol matnini yozing:\n\n"
        "/cancel — bekor qilish",
        parse_mode="HTML")

# poll_step handler - savol va variantlar yig'ish
# Bu admin_msg ichida ishlaydi, alohida state orqali

# ==================== CANCEL ====================

@dp.message(Command("cancel"))
async def cancel_cmd(message: types.Message):
    if message.from_user.id == ADMIN_ID:
        admin_state.pop(ADMIN_ID, None)
        await message.answer("❌ Bekor qilindi. /admin — panelga qaytish")

# ==================== ADMIN XABARLAR ====================

@dp.message(F.from_user.id == ADMIN_ID)
async def admin_msg(message: types.Message):
    # Reply → foydalanuvchiga yo'naltirish
    if message.reply_to_message:
        forward_map = load_json(FORWARD_MAP_FILE)
        target_uid = forward_map.get(str(message.reply_to_message.message_id))
        if target_uid:
            try:
                await bot.copy_message(target_uid, from_chat_id=message.chat.id, message_id=message.message_id)
                await message.answer("✅ Javobingiz foydalanuvchiga yuborildi!")
            except Exception as e:
                await message.answer(f"❌ Yuborilmadi: {e}")
            return

    state = admin_state.get(ADMIN_ID)
    if not state:
        return
    action = state.get("action")

    # ---- OPROS (POLL) YARATISH ----
    poll_step = state.get("poll_step")
    if poll_step:
        poll_data = state.get("poll_data", {})
        txt = message.text.strip() if message.text else ""

        if poll_step == "wait_question":
            if not txt:
                await message.answer("❌ Matn yozing!")
                return
            state["poll_data"]["question"] = txt
            state["poll_step"] = "wait_options"
            admin_state[ADMIN_ID] = state
            await message.answer(
                "✅ Savol saqlandi!\n\n"
                "2-qadam: Javob variantlarini yozing — <b>har biri yangi qatorda</b>:\n\n"
                "Misol:\n<code>Ha\nYo'q\nBilmayman</code>\n\n"
                "Kamida 2 ta, ko'pi bilan 10 ta variant.\n/cancel — bekor qilish",
                parse_mode="HTML")
            return

        elif poll_step == "wait_options":
            options = [o.strip() for o in txt.splitlines() if o.strip()]
            if len(options) < 2:
                await message.answer("❌ Kamida 2 ta variant kiriting!\n\nHar birini yangi qatorda yozing.", parse_mode="HTML")
                return
            if len(options) > 10:
                await message.answer("❌ Ko'pi bilan 10 ta variant bo'lishi mumkin!", parse_mode="HTML")
                return
            state["poll_data"]["options"] = options
            state["poll_step"] = "wait_anonymous"
            admin_state[ADMIN_ID] = state
            await message.answer(
                "✅ Variantlar saqlandi!\n\n"
                "3-qadam: Opros turi:\n\n"
                "<code>1</code> — Anonim (kim ovoz berganini ko'rib bo'lmaydi)\n"
                "<code>2</code> — Ochiq (ovozlar ko'rinadi)\n",
                parse_mode="HTML")
            return

        elif poll_step == "wait_anonymous":
            if txt == "1":
                state["poll_data"]["is_anonymous"] = True
            elif txt == "2":
                state["poll_data"]["is_anonymous"] = False
            else:
                await message.answer("❌ Faqat <code>1</code> yoki <code>2</code> yozing!", parse_mode="HTML")
                return
            state["poll_step"] = None
            admin_state[ADMIN_ID] = state

            # Poll tayyor — endi qaysi action ekanligiga qarab davom etamiz
            poll_data = state["poll_data"]
            anon_text = "Anonim" if poll_data["is_anonymous"] else "Ochiq"
            opts_preview = "\n".join([f"  • {o}" for o in poll_data["options"]])
            await message.answer(
                f"✅ <b>Opros tayyor!</b>\n\n"
                f"❓ <b>{poll_data['question']}</b>\n{opts_preview}\n\n"
                f"🔒 Tur: {anon_text}",
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="📤 Yuborish", callback_data="confirm_poll_send")],
                    [InlineKeyboardButton(text="❌ Bekor", callback_data="back_admin")],
                ])
            )
            return

    # ---- TARIX ID BO'YICHA QIDIRISH ----
    elif action == "history_search_id":
        try:
            target_uid = str(int(message.text.strip()))
        except:
            await message.answer("❌ Faqat raqam kiriting!", parse_mode="HTML")
            return
        admin_state.pop(ADMIN_ID, None)
        history = load_json(HISTORY_FILE)
        users = load_json(USERS_FILE)
        records = history.get(target_uid, [])
        u = users.get(target_uid, {})
        name = f"{u.get('first_name','')} {u.get('last_name','')}".strip() or "Nomsiz"
        username = f"@{u['username']}" if u.get('username') else "—"
        if not records:
            await message.answer(f"❌ <code>{target_uid}</code> ID li foydalanuvchi tarixi topilmadi!", parse_mode="HTML")
            return
        text = (
            f"📋 <b>To'liq tarix</b>\n"
            f"👤 <b>{name}</b> | {username}\n"
            f"🆔 <code>{target_uid}</code>\n"
            f"━━━━━━━━━━━━━━━━\n\n"
        )
        for i, r in enumerate(records, 1):
            event_text = history_event_text(r.get("reason", ""))
            text += f"{i}. {event_text}\n"
            text += f"   🕐 {r.get('event_at', '—')}\n"
            if r.get("phone") and r["phone"] != "admin_qoshgan":
                text += f"   📞 {r['phone']}\n"
            text += "\n"
        back_kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔙 Admin panel", callback_data="back_admin")]
        ])
        for chunk in [text[i:i+4000] for i in range(0, len(text), 4000)]:
            await message.answer(chunk, parse_mode="HTML", reply_markup=back_kb)

    # ---- FOYDALANUVCHI QO'SHISH ----
    if action == "add_user_id":
        try:
            new_uid = str(int(message.text.strip()))
        except:
            await message.answer("❌ Faqat raqam kiriting! Masalan: <code>123456789</code>", parse_mode="HTML")
            return
        users = load_json(USERS_FILE)
        if new_uid in users:
            await message.answer(
                f"⚠️ Bu foydalanuvchi (<code>{new_uid}</code>) allaqachon ro'yxatda bor!",
                parse_mode="HTML")
            admin_state.pop(ADMIN_ID, None)
            return
        users[new_uid] = {
            "id": int(new_uid),
            "first_name": "Qo'shilgan",
            "last_name": "",
            "username": "",
            "phone": "admin_qoshgan",
            "joined": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "left": "",
        }
        save_json(USERS_FILE, users)
        admin_state.pop(ADMIN_ID, None)
        await message.answer(
            f"✅ Foydalanuvchi <code>{new_uid}</code> ro'yxatga qo'shildi!\n"
            f"Endi unga rejalik postlar yuboriladi.",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔙 Admin panel", callback_data="back_admin")]
            ])
        )
        await schedule_posts_for_user(int(new_uid))

    # ---- FOYDALANUVCHINI BOSHQARISH (ID KIRITISH) ----
    elif action == "manage_user_id":
        try:
            target_uid = str(int(message.text.strip()))
        except:
            await message.answer("❌ Faqat raqam kiriting!", parse_mode="HTML")
            return
        users = load_json(USERS_FILE)
        blocked = load_json(BLOCKED_FILE)
        if target_uid not in users:
            await message.answer(f"❌ <code>{target_uid}</code> ID li foydalanuvchi topilmadi!", parse_mode="HTML")
            admin_state.pop(ADMIN_ID, None)
            return
        u = users[target_uid]
        name = f"{u.get('first_name','')} {u.get('last_name','')}".strip() or "Nomsiz"
        username = f"@{u['username']}" if u.get('username') else "—"
        phone = u.get('phone') or "—"
        is_block = target_uid in blocked
        status = "🚫 Bloklangan" if is_block else ("❌ Botdan chiqqan" if u.get('left') else "✅ Faol")
        admin_state.pop(ADMIN_ID, None)
        await message.answer(
            f"👤 <b>{name}</b>\n"
            f"🆔 <code>{target_uid}</code>\n"
            f"👤 {username} | 📞 {phone}\n"
            f"📅 Qo'shilgan: {u.get('joined','—')}\n"
            f"🔰 Holati: {status}",
            parse_mode="HTML",
            reply_markup=user_manage_keyboard(target_uid, is_block)
        )

    # ---- ODDIY POST ----
    elif action == "broadcast_now":
        users = load_json(USERS_FILE)
        success = failed = 0
        await message.answer(f"📤 Yuborilmoqda... ({len(users)} ta a'zo)")
        for uid in users.keys():
            if is_blocked(int(uid)):
                continue
            try:
                await message.copy_to(int(uid))
                success += 1
                await asyncio.sleep(0.05)
            except:
                failed += 1
        await message.answer(f"✅ Yuborildi!\n✅ {success} ta | ❌ {failed} ta")
        admin_state[ADMIN_ID] = {"action": "broadcast_now_feedback_delay"}
        await message.answer(
            "🎬 <b>«Ko'rdingizmi?» xabarini qachon yuborish kerak?</b>\n\n"
            "Misol:\n"
            "<code>0</code> — darhol\n"
            "<code>10 minut</code> — 10 minutdan keyin\n"
            "<code>2 soat</code> — 2 soatdan keyin\n"
            "<code>1 kun</code> — 1 kundan keyin\n\n"
            "/skip — «Ko'rdingizmi?» yubormaslik",
            parse_mode="HTML"
        )

    elif action == "timed_wait_post":
        post_data = {}
        if message.photo:
            post_data["photo_id"] = message.photo[-1].file_id
            post_data["text"] = message.caption or ""
        elif message.video:
            post_data["video_id"] = message.video.file_id
            post_data["text"] = message.caption or ""
        elif message.video_note:
            post_data["video_note_id"] = message.video_note.file_id
        elif message.text:
            post_data["text"] = message.text
        else:
            await message.answer("❌ Faqat rasm, video, dumaloq video yoki matn!")
            return
        admin_state[ADMIN_ID] = {"action": "timed_wait_time", "post_data": post_data}
        await message.answer("✅ Post qabul qilindi!\n\n⏰ Vaqtni yozing:\n<b>Format: YYYY-MM-DD HH:MM</b>\nMasalan: <code>2025-06-10 14:30</code>", parse_mode="HTML")

    elif action == "timed_wait_time":
        try:
            send_time = datetime.strptime(message.text.strip(), "%Y-%m-%d %H:%M")
            if send_time <= datetime.now():
                await message.answer("⚠️ Vaqt o'tib ketgan! Kelajakdagi vaqt kiriting.", parse_mode="HTML")
                return
            post_data = state["post_data"]
            post_data["send_time"] = send_time.strftime("%Y-%m-%d %H:%M")
            post_data["sent"] = False
            timed = load_json(TIMED_FILE)
            pid = str(int(datetime.now().timestamp()))
            timed[pid] = post_data
            save_json(TIMED_FILE, timed)
            await message.answer(f"✅ Post rejalashtirildi!\n📅 Vaqt: <b>{send_time.strftime('%Y-%m-%d %H:%M')}</b>", parse_mode="HTML")
            admin_state[ADMIN_ID] = {"action": "timed_feedback_delay", "post_send_time": send_time.strftime("%Y-%m-%d %H:%M")}
            await message.answer(
                "🎬 <b>«Ko'rdingizmi?» xabarini postdan necha vaqt keyin yuborish kerak?</b>\n\n"
                "Misol:\n"
                "<code>0</code> — postdan darhol keyin\n"
                "<code>10 minut</code>\n"
                "<code>2 soat</code>\n"
                "<code>1 kun</code>\n\n"
                "/skip — «Ko'rdingizmi?» yubormaslik",
                parse_mode="HTML"
            )
        except ValueError:
            await message.answer("❌ Noto'g'ri format!\nMasalan: <code>2025-06-10 14:30</code>", parse_mode="HTML")

    # ---- BROADCAST_NOW FEEDBACK DELAY ----
    elif action == "broadcast_now_feedback_delay":
        txt = message.text.strip().lower() if message.text else ""
        try:
            if txt in ("0", "darhol"):
                delay_hours = 0
            elif "minut" in txt:
                delay_hours = float(txt.replace("minut", "").strip()) / 60
            elif "soat" in txt:
                delay_hours = float(txt.replace("soat", "").strip())
            elif "kun" in txt:
                delay_hours = float(txt.replace("kun", "").strip()) * 24
            else:
                delay_hours = float(txt)
        except:
            await message.answer("❌ Noto'g'ri format!\nMisol: <code>10 minut</code>, <code>2 soat</code>, <code>1 kun</code>", parse_mode="HTML")
            return
        send_time = datetime.now() + timedelta(hours=delay_hours)
        ref = f"feedback_{int(datetime.now().timestamp())}"
        timed = load_json(TIMED_FILE)
        pid = str(int(datetime.now().timestamp())) + "_fb"
        timed[pid] = {
            "feedback": True,
            "ref": ref,
            "send_time": send_time.strftime("%Y-%m-%d %H:%M"),
            "sent": False
        }
        save_json(TIMED_FILE, timed)
        admin_state.pop(ADMIN_ID, None)
        delay_text = format_delay(delay_hours)
        await message.answer(
            f"✅ «Ko'rdingizmi?» xabari <b>{delay_text}</b> dan keyin yuboriladi.\n"
            f"📅 Vaqt: <b>{send_time.strftime('%Y-%m-%d %H:%M')}</b>",
            parse_mode="HTML",
            reply_markup=next_step_keyboard()
        )

    # ---- TIMED POST FEEDBACK DELAY ----
    elif action == "timed_feedback_delay":
        txt = message.text.strip().lower() if message.text else ""
        post_send_time_str = state.get("post_send_time", "")
        try:
            if txt in ("0", "darhol"):
                delay_hours = 0
            elif "minut" in txt:
                delay_hours = float(txt.replace("minut", "").strip()) / 60
            elif "soat" in txt:
                delay_hours = float(txt.replace("soat", "").strip())
            elif "kun" in txt:
                delay_hours = float(txt.replace("kun", "").strip()) * 24
            else:
                delay_hours = float(txt)
        except:
            await message.answer("❌ Noto'g'ri format!\nMisol: <code>10 minut</code>, <code>2 soat</code>, <code>1 kun</code>", parse_mode="HTML")
            return
        base_time = datetime.strptime(post_send_time_str, "%Y-%m-%d %H:%M") if post_send_time_str else datetime.now()
        send_time = base_time + timedelta(hours=delay_hours)
        ref = f"feedback_{int(datetime.now().timestamp())}"
        timed = load_json(TIMED_FILE)
        pid = str(int(datetime.now().timestamp())) + "_fb"
        timed[pid] = {
            "feedback": True,
            "ref": ref,
            "send_time": send_time.strftime("%Y-%m-%d %H:%M"),
            "sent": False
        }
        save_json(TIMED_FILE, timed)
        admin_state.pop(ADMIN_ID, None)
        delay_text = format_delay(delay_hours)
        await message.answer(
            f"✅ «Ko'rdingizmi?» xabari postdan <b>{delay_text}</b> keyin yuboriladi.\n"
            f"📅 Vaqt: <b>{send_time.strftime('%Y-%m-%d %H:%M')}</b>",
            parse_mode="HTML",
            reply_markup=next_step_keyboard()
        )

    elif action == "set_welcome":
        posts = load_json(POSTS_FILE)
        if "welcome" not in posts:
            posts["welcome"] = {}
        if message.voice:
            posts["welcome"]["voice_id"] = message.voice.file_id
            save_json(POSTS_FILE, posts)
            admin_state.pop(ADMIN_ID, None)
            await message.answer("✅ Ovozli xabar saqlandi!")
        elif message.photo:
            posts["welcome"]["photo_id"] = message.photo[-1].file_id
            posts["welcome"]["text"] = message.caption or ""
            save_json(POSTS_FILE, posts)
            await message.answer("✅ Rasm saqlandi!\n\nOvozli xabar ham qo'shmoqchimisiz? Yuboring yoki /cancel:")
        elif message.text:
            posts["welcome"]["text"] = message.text
            save_json(POSTS_FILE, posts)
            admin_state.pop(ADMIN_ID, None)
            await message.answer("✅ Matn saqlandi!")

    elif action == "add_post_wait_content":
        post_data = {}
        if message.photo:
            post_data["photo_id"] = message.photo[-1].file_id
            post_data["text"] = message.caption or ""
        elif message.video:
            post_data["video_id"] = message.video.file_id
            post_data["text"] = message.caption or ""
        elif message.video_note:
            post_data["video_note_id"] = message.video_note.file_id
        elif message.text:
            post_data["text"] = message.text
        else:
            await message.answer("❌ Faqat rasm, video, dumaloq video yoki matn!")
            return
        admin_state[ADMIN_ID] = {"action": "add_post_wait_delay", "post_data": post_data}
        await message.answer(
            "✅ Post qabul qilindi!\n\n⏰ Necha vaqt keyin yuborilsin?\n\n"
            "<code>0</code> — darhol\n<code>10 minut</code>\n<code>5 soat</code>\n<code>3 kun</code>",
            parse_mode="HTML")

    elif action == "add_post_wait_delay":
        text = message.text.strip().lower()
        try:
            if text in ("0", "darhol"):
                delay_hours = 0
            elif "minut" in text:
                delay_hours = float(text.replace("minut", "").strip()) / 60
            elif "soat" in text:
                delay_hours = float(text.replace("soat", "").strip())
            elif "kun" in text:
                delay_hours = float(text.replace("kun", "").strip()) * 24
            else:
                delay_hours = float(text)
        except:
            await message.answer("❌ Noto'g'ri format!\nMisol: <code>10 minut</code>, <code>5 soat</code>, <code>3 kun</code>", parse_mode="HTML")
            return
        post_data = state["post_data"]
        post_data["delay_hours"] = delay_hours
        posts = load_json(POSTS_FILE)
        existing = [k for k in posts.keys() if k.startswith("post_")]
        post_key = f"post_{len(existing) + 1}"
        posts[post_key] = post_data
        save_json(POSTS_FILE, posts)
        admin_state.pop(ADMIN_ID, None)
        await message.answer(
            f"✅ <b>{post_key}</b> saqlandi!\n⏰ Vaqt: <b>{format_delay(delay_hours)}</b>",
            parse_mode="HTML", reply_markup=next_step_keyboard())

    elif action == "update_post":
        post_key = state["post_key"]
        posts = load_json(POSTS_FILE)
        old_post = posts.get(post_key, {})
        if message.voice:
            old_post["voice_id"] = message.voice.file_id
            posts[post_key] = old_post
            save_json(POSTS_FILE, posts)
            admin_state.pop(ADMIN_ID, None)
            await message.answer(f"✅ {post_key} ovozli xabari yangilandi!", reply_markup=next_step_keyboard())
        elif message.photo:
            old_post["photo_id"] = message.photo[-1].file_id
            old_post["text"] = message.caption or ""
            old_post.pop("video_id", None)
            old_post.pop("video_note_id", None)
            posts[post_key] = old_post
            save_json(POSTS_FILE, posts)
            await message.answer("✅ Rasm yangilandi!\n\nOvozli xabar ham qo'shmoqchimisiz? Yoki /cancel:")
        elif message.video:
            old_post["video_id"] = message.video.file_id
            old_post["text"] = message.caption or ""
            old_post.pop("photo_id", None)
            old_post.pop("video_note_id", None)
            posts[post_key] = old_post
            save_json(POSTS_FILE, posts)
            admin_state.pop(ADMIN_ID, None)
            await message.answer(f"✅ {post_key} yangilandi!", reply_markup=next_step_keyboard())
        elif message.video_note:
            old_post["video_note_id"] = message.video_note.file_id
            old_post.pop("photo_id", None)
            old_post.pop("video_id", None)
            posts[post_key] = old_post
            save_json(POSTS_FILE, posts)
            admin_state.pop(ADMIN_ID, None)
            await message.answer(f"✅ {post_key} dumaloq video yangilandi!", reply_markup=next_step_keyboard())
        elif message.text:
            old_post["text"] = message.text
            posts[post_key] = old_post
            save_json(POSTS_FILE, posts)
            admin_state.pop(ADMIN_ID, None)
            await message.answer(f"✅ {post_key} matni yangilandi!", reply_markup=next_step_keyboard())

# ==================== BOTDAN CHIQISH ====================

@dp.my_chat_member()
async def user_left(event: types.ChatMemberUpdated):
    if event.new_chat_member.status in ["kicked", "left"]:
        users = load_json(USERS_FILE)
        uid = str(event.from_user.id)
        if uid in users:
            users[uid]["left"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            save_json(USERS_FILE, users)
            save_history(uid, users[uid], reason="self_left", by="self")

# ==================== ODDIY XABARLAR ====================

@dp.message()
async def any_msg(message: types.Message):
    if message.from_user.id == ADMIN_ID:
        return
    if is_blocked(message.from_user.id):
        await message.answer("⛔ Siz botdan bloklangansiz.")
        return
    is_sub = await check_subscription(message.from_user.id)
    if not is_sub:
        await message.answer("⚠️ Botdan foydalanish uchun kanalga obuna bo'ling!", reply_markup=sub_keyboard())
        return
    await message.answer("📩 Savolingiz bo'lsa, @Mehriddin_Nuriddinov1 adminiga murojaat qiling!")

# ==================== ISHGA TUSHIRISH ====================

async def main():
    logger.info("Bot ishga tushdi!")
    asyncio.create_task(check_and_send_scheduled())
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
